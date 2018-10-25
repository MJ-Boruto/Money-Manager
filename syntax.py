from openpyxl import Workbook
wb = Workbook()  #創立EXCEL檔，創立的excel會在和.py檔同一個資料夾
ws = wb.active #開啟分頁簿
import datetime #日期自動產生
today = datetime.date.today() 

ws['A1'] = '日期'
ws['B1'] = '收支'
ws['C1'] = '類別'
ws['D1'] = '金額'
ws['E1'] = '餘額'
#輸入excel資料

ws['F1'] = '支出項目'
ws['F2'] = '1.用餐'
ws['F3'] = '2.飲品'
ws['F4'] = '3.娛樂'
ws['F5'] = '總支出'
ws['G1'] = '金額'
#統計表表格資訊

total=0 #餘額
dine=0 #用餐總花費
drink=0 #飲品總花費
play=0 #娛樂總花類
print('歡迎來到記帳程式\n本程式將幫助您紀錄日常收支\n包刮收入、支出(用餐、飲品、娛樂)\n紀錄將自動生成Excel檔「記帳.xlsx」(紀錄中請勿開啟該Excel檔)\n請依照以下順序填寫紀錄：\n')
for row in range(2, 100):
	ws.cell(column=1, row=row).value=today #自動輸入日期
	gol=input('請選擇收入或支出\n1.收入\n2.支出\n')
	#先選支出或收入(gol=gain or lose, 收入不選項目)
	if gol=='1':
		ws.cell(column=2, row=row).value='收入'
		ws.cell(column=3, row=row).value='-'
		revenue=int(input('請問收入多少？\n'))
		ws.cell(column=4, row=row).value=revenue
		total+=revenue
	if gol=='2':
		ws.cell(column=2, row=row).value='支出'
		kind=input('請選擇消費種類\n1.用餐\n2.飲品\n3.娛樂\n')
		#再選支出項目
		if kind=='1':
			ws.cell(column=3, row=row).value='用餐'
			expense=-int(input('用餐花費多少？\n'))
			ws.cell(column=4, row=row).value=expense
			total+=expense
			dine+=expense
		if kind=='2':
			ws.cell(column=3, row=row).value='飲品'
			expense=-int(input('飲品花費多少？\n'))
			ws.cell(column=4, row=row).value=expense
			total+=expense
			drink+=expense
		if kind=='3':
			ws.cell(column=3, row=row).value='娛樂'
			expense=-int(input('娛樂花費多少？\n'))
			ws.cell(column=4, row=row).value=expense
			total+=expense
			play+=expense
	ws.cell(column=5, row=row).value=total #餘額
	if gol=='1': #有點難解釋，你比較填「收入」和「支出」時跑出來資訊的差別就知道了
		print('今日'+str(today)+'，獲得收入'+str(revenue)+'，本次餘額'+str(total))
	if gol=='2':
		print('今日'+str(today)+'，支出金額'+str(expense)+'，本次餘額'+str(total)+'\n支出項目：\n1.用餐'+'$'+str(dine)+'元'+'\n2.飲品'+'$'+str(drink)+'元'+'\n3.娛樂'+'$'+str(play)+'元''\n總花費金額'+'$'+str(dine+drink+play)+'元')

	ws['G2'] = dine 
	ws['G3'] = drink
	ws['G4'] = play
	consumption=dine+drink+play
	ws['G5'] = consumption
	#匯入統計資訊
	wb.save("記帳.xlsx")
	#最重要的步驟-存檔
	#記得跑程式時excel檔要關掉，不然不能存取
