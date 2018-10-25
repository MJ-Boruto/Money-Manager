import tkinter as tk
from tkinter import ttk
import os
import openpyxl
import datetime
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg, NavigationToolbar2Tk
from matplotlib.figure import Figure
from openpyxl.styles import Alignment #這樣才能置中

LARGE_FONT = ("微軟正黑體", 14)
SMALL_FONT = ("微軟正黑體", 10)

if os.path.exists('記帳.xlsx') != True:


			wb = openpyxl.Workbook()
			ws = wb['Sheet']
			ws.title ='明細資料'
			ws['A1'] = '日期'
			ws['B1'] = '收支'
			ws['C1'] = '類別'
			ws['D1'] = '金額'
			ws['E1'] = '餘額'
			ws['F1'] = 0
			wb.create_sheet(title='統計資料')

			#以下為統計數據的欄位設置
			ws2 = wb['統計資料']
			ws2.merge_cells('A1:D1') #合併欄位
			
			align = Alignment(horizontal='center', vertical='center') #先定義樣式"置中"
			ws2['A1'].alignment = align #然後就置中了XD


			ws2['A1'] = '收支項目'
			ws2.merge_cells('A2:B2')

			ws2['A2'].alignment = align


			ws2['A2'] = '收入'
			ws2['A3'] = '1.收入-薪水'
			ws2['A4'] = '2.收入-兼職'
			ws2['A5'] = '3.收入-其他'
			ws2['A8'] = '總收入'
			ws2.merge_cells('C2:D2')

			#讓每格有數字，以利計算
			ws2['B3'] = 0
			ws2['B4'] = 0
			ws2['B5'] = 0
			ws2['B6'] = 0
			ws2['C2'].alignment = align
			ws2['C2'] = '支出'
			ws2['C3'] = '1.支出-飲食'
			ws2['C4'] = '2.支出-交通'
			ws2['C5'] = '3.支出-居家'
			ws2['C6'] = '4.支出-娛樂'
			ws2['C7'] = '5.支出-其他'
			ws2['C8'] = '總支出'
			ws2['D3'] = 0
			ws2['D4'] = 0
			ws2['D5'] = 0
			ws2['D6'] = 0
			ws2['D7'] = 0
			ws2['D8'] = 0

			wb.save('記帳.xlsx')
wb = openpyxl.load_workbook('記帳.xlsx', data_only=True)

class App(tk.Tk):

	def __init__(self, *arg, **kwargs):
		
		tk.Tk.__init__(self, *arg, **kwargs)
		tk.Tk.wm_title(self, "記帳小幫手><")

		container = tk.Frame(self)
		container.pack(side="top", fill="both", expand = True)
		'''#抄來的code原本有，但沒有好像也沒差
		container.grid_rowconfigure(0, weight=1)
		container.grid_columnconfigure(0, weight=1)
		'''
		self.frames = {}

		for F in (StartPage, RecordPage, PieChart, BarChart):
			frame = F(container, self)
			self.frames[F] = frame
			frame.grid(row=0, column=0, sticky="nsew")
		
		self.show_frame(StartPage)

	def show_frame(self, cont):
		frame = self.frames[cont]
		frame.tkraise()

class StartPage(tk.Frame):
	
	def __init__(self, parent, controller):
		tk.Frame.__init__(self, parent)
		label = tk.Label(self, text="記帳小幫手><", font=LARGE_FONT)
		label2 = tk.Label(self, text="\n統計報表：")
		button = ttk.Button(self, text="記一筆",
			command=lambda: controller.show_frame(RecordPage))
		button2 = ttk.Button(self, text="分類圓餅圖",
			command=lambda: controller.show_frame(PieChart))
		button3 = ttk.Button(self, text="月長條圖",
			command=lambda: controller.show_frame(BarChart))
		label.pack(pady=10,padx=10)
		button.pack()
		label2.pack()
		button2.pack()
		button3.pack()

class RecordPage(tk.Frame):

	
	def __init__(self, parent, controller):
		tk.Frame.__init__(self, parent)
		label = tk.Label(self, text="新增收支紀錄", font=LARGE_FONT)
		label.grid(row=0,column=1,pady=20)

		label2 = tk.Label(self, text="日期：", font=SMALL_FONT)
		label2.grid(row=1,column=0,sticky="e")
		entry_date = tk.Entry(self)
		now = datetime.datetime.now()
		entry_date.insert(0,'{}/{}/{}'.format(now.year,now.month,now.day))
		entry_date.grid(row=1,column=1,columnspan=2,sticky="w")

		label3 = tk.Label(self, text="收支：", font=SMALL_FONT)
		label3.grid(row=2,column=0,sticky="e")
		combobox = ttk.Combobox(self, state="readonly",value=['支出','收入'])
		combobox.current(0)
		combobox.grid(row=2,column=1,columnspan=2,sticky="w")
		combobox.bind('<<ComboboxSelected>>', lambda x: set_type())

		label4 = tk.Label(self, text="類別：", font=SMALL_FONT)
		label4.grid(row=3,column=0,sticky="e")
		combobox2 = ttk.Combobox(self, state="readonly")
		combobox2['value']='飲食','交通','居家','娛樂','其他'
		combobox2.current(0)
		combobox2.grid(row=3,column=1,columnspan=2,sticky="w")

		label6 = tk.Label(self, text="金額：", font=SMALL_FONT)
		label6.grid(row=4,column=0,sticky="e")
		entry = tk.Entry(self)
		entry.grid(row=4,column=1,columnspan=2,sticky="w")
		entry.insert(0,0)

		button3 = ttk.Button(self, text="<返回",
			command=lambda: controller.show_frame(StartPage))
		button3.grid(row=5,column=0,pady=15)
		button4 = ttk.Button(self, text="再記一筆",
			command=lambda: record())
		button4.grid(row=5,column=1)
		button5 = ttk.Button(self, text="儲存",
			command=lambda: [controller.show_frame(StartPage),record])
		button5.grid(row=5,column=2)

		def set_type():
			nonlocal combobox
			nonlocal combobox2
			if combobox.get()=='支出':
				combobox2['value']='飲食','交通','居家','娛樂','其他'
				combobox2.current(0)
			else :
				combobox2['value']='薪水','兼職','其他'
				combobox2.current(0)

		def record():
			ws = wb['明細資料']
			ws2 = wb['統計資料']
			row = ws.max_row
			ws['A'+str(row+1)]=entry_date.get()
			ws['B'+str(row+1)]=combobox.get()
			ws['C'+str(row+1)]=combobox2.get()

			if combobox.get()=='收入':
				ws['D'+str(row+1)]=entry.get()
				#判斷如果是該項目，就加到"統計資料裡"
				if combobox2.get()=='薪水':	
					ws2['B3'] = str(int(ws2['B3'].value)+ int(ws['D'+str(row+1)].value))
				elif combobox2.get()=='兼職':
					ws2['B4'] = str(int(ws2['B4'].value)+ int(ws['D'+str(row+1)].value))
				elif combobox2.get()=='其他':
					ws2['B5'] = str(int(ws2['B5'].value)+ int(ws['D'+str(row+1)].value))
				#分隔線
			else:
				ws['D'+str(row+1)]=str(-1*int(entry.get()))
				#判斷如果是該項目，就加到"統計資料裡"
				if combobox2.get()=='飲食':	
					ws2['D3'] = str(int(ws2['D3'].value)+ int(ws['D'+str(row+1)].value))
				elif combobox2.get()=='交通':
					ws2['D4'] = str(int(ws2['D4'].value)+ int(ws['D'+str(row+1)].value))
				elif combobox2.get()=='居家':
					ws2['D5'] = str(int(ws2['D5'].value)+ int(ws['D'+str(row+1)].value))
				elif combobox2.get()=='娛樂':
					ws2['D6'] = str(int(ws2['D6'].value)+ int(ws['D'+str(row+1)].value))
				elif combobox2.get()=='其他':
					ws2['D7'] = str(int(ws2['D7'].value)+ int(ws['D'+str(row+1)].value))
				#分隔線

			ws['F'+str(row+1)]=str(int(ws['F'+str(row)].value)+int(ws['D'+str(row+1)].value))

			#總支出和總收入
			totalR = int(ws2['B3'].value) + int(ws2['B4'].value) + int(ws2['B5'].value)
			totalE = int(ws2['D3'].value) + int(ws2['D4'].value) + int(ws2['D5'].value) + int(ws2['D6'].value) + int(ws2['D7'].value)
			ws2['B8'] = totalR
			ws2['D8'] = totalE


			wb.save('記帳.xlsx')
	
class PieChart(tk.Frame):
	
	ws = wb['統計資料']

	def __init__(self, parent, controller):
		tk.Frame.__init__(self, parent)
		label = tk.Label(self, text="分類圓餅圖", font=LARGE_FONT)
		label.pack()

		button1 = ttk.Button(self, text="<返回",
			command=lambda: controller.show_frame(StartPage))
		button1.pack()

		f = Figure(figsize=(5,5), dpi=50)
		a = f.add_subplot(111)
		a.plot([1,2,3,4,5,6,7,8,9,10,11,12],[5,6,1,3,8,9,3,5,5,5,5,5])

		canvas = FigureCanvasTkAgg(f, self)
		canvas.draw()
		canvas.get_tk_widget().pack(side="top", fill="both", expand = True)

class BarChart(tk.Frame):
	
	ws = wb['統計資料']

	def __init__(self, parent, controller):
		tk.Frame.__init__(self, parent)
		label = tk.Label(self, text="月長條圖", font=LARGE_FONT)
		label.pack()

		button1 = ttk.Button(self, text="<返回",
			command=lambda: controller.show_frame(StartPage))
		button1.pack()

		f = Figure(figsize=(5,5), dpi=50)
		a = f.add_subplot(111)
		a.plot([1,2,3,4,5,6,7,8,9,10,11,12],[5,6,1,3,8,9,3,5,5,5,5,5])

		canvas = FigureCanvasTkAgg(f, self)
		canvas.draw()
		canvas.get_tk_widget().pack(side="top", fill="both", expand = True)

app = App()
app.mainloop()
