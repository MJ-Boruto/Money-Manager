from openpyxl import Workbook
wb = Workbook()
ws = wb.active
#創立一個EXCEL檔，產生的excel會在和.py檔同一個資料夾

ws['A1'] = '日期'
ws['B1'] = '收支'
ws['C1'] = '類別'
ws['D1'] = '金額'
ws['E1'] = '餘額'
#輸入資料

ws['F1'] = '支出項目'
ws['F2'] = '1.用餐'
ws['F3'] = '2.飲品'
ws['F4'] = '3.娛樂'
ws['G1'] = '金額'
#統計表表格資訊

consumption=[]
total=0
dine=0
drink=0
play=0
#因為還沒有辦法即時輸入，先在Python裡個別記下來

#以下單純是把為「記帳.py」的語法複製過來
consumption=[]
total=0
#日期、選擇記帳或報表
#表示這個月的帳款加總，及時存取
while True:
	gol=input('請選擇收入或支出\n1.收入\n2.支出\n')
	#先選支出或收入(gol=gain or lose, 收入不選項目)
	if gol=='1':
		revenue=int(input('請問收入多少？\n'))
		consumption.append(revenue)
		total+=revenue
	if gol=='2':
		kind=input('請選擇消費種類\n1.用餐\n2.飲品\n3.娛樂\n')
		#再選支出項目
		if kind=='1':
			amount=-int(input('用餐花費多少？\n'))
			consumption.append(amount)
			total+=amount
			dine+=amount
		if kind=='2':
			amount=-int(input('飲品花費多少？\n'))
			consumption.append(amount)
			total+=amount
			drink+=amount
		if kind=='3':
			amount=-int(input('娛樂花費多少？\n'))
			consumption.append(amount)
			total+=amount
			play+=amount
	date=input('請輸入日期(例：2018/5/20)\n')
	if gol=='1':#這邊分類可有可無，如果填寫的資訊是支出項目的話會多顯示統計項目，填收入的話就沒有。
		print('今日'+str(date)+'，獲得收入'+str(revenue)+'，本次餘額'+str(total))
	if gol=='2':
		print('今日'+str(date)+'，支出金額'+str(amount)+'，本次餘額'+str(total)+'\n支出項目：\n1.用餐'+'$'+str(dine)+'元'+'\n2.飲品'+'$'+str(drink)+'元'+'\n3.娛樂'+'$'+str(play)+'元''\n總花費金額'+'$'+str(dine+drink+play)+'元')
	ws['G2'] = dine 
	ws['G3'] = drink
	ws['G4'] = play
	wb.save("記帳.xlsx")#儲存，檔案名稱看你要什麼都可以
#以上單純是把為「記帳.py」的語法複製過來

#我想把帳簿的資料即時輸入excel檔案，但我不知道怎麼讓程式在每一筆後自動換到下一列


